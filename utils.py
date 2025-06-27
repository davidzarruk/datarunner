import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageOps, ImageDraw, ImageFont, ImageOps
import io
import unicodedata
from PIL import Image, ImageDraw
import os
import boto3
from io import BytesIO


def upload_df_to_s3(df, bucket_name, key):
    s3 = boto3.client('s3')
    buffer = BytesIO()

    df.to_csv(buffer, index=False, encoding='utf-8')
    content_type = 'text/csv'

    buffer.seek(0)  # Rewind the buffer to the beginning
    s3.put_object(Bucket=bucket_name, Key=key, Body=buffer, ContentType=content_type)


def read_df_from_s3(bucket_name, key):
    s3 = boto3.client('s3')
    response = s3.get_object(Bucket=bucket_name, Key=key)
    body = response['Body'].read()

    df = pd.read_csv(BytesIO(body), encoding='utf-8')

    return df


TABLE_OPACITY = 255
FONT = "HelveticaNeueMedium.otf"

def remove_accents(text):
    return ''.join(
        c for c in unicodedata.normalize('NFKD', text)
        if not unicodedata.combining(c)
    )



country_map = {
    'CHI': 'Chile',
    'CHL': 'Chile',
    'ECU': 'Ecuador',
    'ARG': 'Argentina',
    'COL': 'Colombia',
    'BRA': 'Brasil',
    'PER': 'Perú',
    'PAR': 'Paraguay',
    'PRY': 'Paraguay',
    'CRC': 'Costa Rica',
    'CRI': 'Costa Rica',
    'BOL': 'Bolivia',
    'MEX': 'México',
    'SLV': 'El Salvador',
    'ESA': 'El Salvador',
    'GTM': 'Guatemala',
    'GUA': 'Guatemala',
    'VEN': 'Venezuela',
    'URU': 'Uruguay',
    'URY': 'Uruguay',
    'PAN': 'Panamá',
    'HON': 'Honduras',
    'HND': 'Honduras',
    'DOM': 'Rep. Dominicana',
    'CUB': 'Cuba',
    'HAI': 'Haiti',
    'CL': 'Chile',
    'EC': 'Ecuador',
    'AR': 'Argentina',
    'CO': 'Colombia',
    'BR': 'Brasil',
    'PE': 'Perú',
    'PY': 'Paraguay',
    'CR': 'Costa Rica',
    'BO': 'Bolivia',
    'MX': 'México',
    'SV': 'El Salvador',
    'GT': 'Guatemala',
    'VE': 'Venezuela',
    'UY': 'Uruguay',
    'PA': 'Panamá',
    'HN': 'Honduras',
    'DO': 'Rep. Dominicana',
    'CU': 'Cuba',
    'HT': 'Haiti'}


categories = {
    'JU20': 'Menos de 20',
    'H': '20-29',
    '30': '30-34',
    '35': '35-39',
    '40': '40-44',
    '45': '45-49',
    '50': '50-54',
    '55': '55-59',
    '60': '60-64',
    '65': '65-69',
    '70': '70-74',
    '75': '75-79',
    '80': '80 o más'
}


categories = {
    'Abierta': '18-39',
    'Master': '40-49',
    'Plus': '50 o más'
}


boston_qualifying_times = {
    '\x13': {'M': '2:00:00', 'W': '2:00:00'},
    'JU20': {'M': '2:55:00', 'W': '3:25:00'},
    'H': {'M': '2:55:00', 'W': '3:25:00'},
    '30': {'M': '2:55:00', 'W': '3:25:00'},
    '35': {'M': '3:00:00', 'W': '3:30:00'},
    '40': {'M': '3:05:00', 'W': '3:35:00'},
    '45': {'M': '3:15:00', 'W': '3:45:00'},
    '50': {'M': '3:20:00', 'W': '3:50:00'},
    '55': {'M': '3:30:00', 'W': '4:00:00'},
    '60': {'M': '3:50:00', 'W': '4:20:00'},
    '65': {'M': '4:05:00', 'W': '4:35:00'},
    '70': {'M': '4:20:00', 'W': '4:50:00'},
    '75': {'M': '4:35:00', 'W': '5:05:00'},
    '80': {'M': '4:50:00', 'W': '5:20:00'}
}

boston_qualifying_times = {
    'Abierta': {'M': '2:55:00', 'W': '3:25:00'},
    'Master': {'M': '3:05:00', 'W': '3:35:00'},
    'Plus': {'M': '3:20:00', 'W': '3:50:00'},
}


# Country code mapping dictionary
country_mapping_cleaning = {
    'GER': 'DEU', 'DEU': 'DEU',
    'GBR': 'GBR', 'ENG': 'GBR', 'SCO': 'GBR', 'WLS': 'GBR', 'NIR': 'GBR',
    'USA': 'USA',
    'DNK': 'DNK', 'DEN': 'DNK', 'DK': 'DNK',
    'FRA': 'FRA',
    'ITA': 'ITA',
    'ESP': 'ESP',
    'SWE': 'SWE',
    'NLD': 'NLD', 'NED': 'NLD', 'HOL': 'NLD',
    'AUT': 'AUT',
    'BRA': 'BRA',
    'POL': 'POL',
    'MEX': 'MEX',
    'NOR': 'NOR',
    'FIN': 'FIN',
    'CHE': 'CHE', 'SUI': 'CHE',
    'CHN': 'CHN',
    'BEL': 'BEL',
    'IRL': 'IRL',
    'CAN': 'CAN',
    'AUS': 'AUS',
    'JPN': 'JPN', 'JAP': 'JPN',
    'RUS': 'RUS',
    'INA': 'IDN', 'IDN': 'IDN',
    'COL': 'COL',
    'ISR': 'ISR',
    'HKG': 'HKG',
    'ARG': 'ARG',
    'IND': 'IND',
    'THA': 'THA',
    'VEN': 'VEN',
    'TWN': 'TWN', 'TPE': 'TWN',
    'RSA': 'ZAF', 'ZAF': 'ZAF',
    'UKR': 'UKR',
    'CRC': 'CRI', 'CRI': 'CRI',
    'KOR': 'KOR',
    'HUN': 'HUN',
    'POR': 'PRT', 'PRT': 'PRT',
    'CHI': 'CHL', 'CHL': 'CHL',
    'NZL': 'NZL',
    'CZE': 'CZE',
    'PHI': 'PHL', 'PHL': 'PHL',
    'EST': 'EST',
    'TUR': 'TUR',
    'SVN': 'SVN',
    'PER': 'PER',
    'ISL': 'ISL',
    'LUX': 'LUX',
    'ROU': 'ROU', 'ROM': 'ROU',
    'SVK': 'SVK',
    'GUA': 'GTM', 'GTM': 'GTM',
    'ECU': 'ECU',
    'GRE': 'GRC', 'GRC': 'GRC',
    'DOM': 'DOM',
    'MAR': 'MAR',
    'MAS': 'MYS', 'MYS': 'MYS',
    'LTU': 'LTU', 'LIT': 'LTU',
    'CRO': 'HRV', 'HRV': 'HRV',
    'PUR': 'PRI', 'PRI': 'PRI',
    'SIN': 'SGP', 'SGP': 'SGP',
    'SRB': 'SRB',
    'KEN': 'KEN',
    'KAZ': 'KAZ',
    'LAT': 'LVA', 'LVA': 'LVA',
    'PAN': 'PAN',
    'BLR': 'BLR',
    'BUL': 'BGR', 'BGR': 'BGR',
    'ETH': 'ETH',
    'URY': 'URY',
    'EGY': 'EGY',
    'CYP': 'CYP',
    'FRO': 'FRO',
    'PAR': 'PRY', 'PRY': 'PRY',
    'IRI': 'IRN', 'IRN': 'IRN',
    'HON': 'HND', 'HND': 'HND',
    'VIE': 'VNM', 'VNM': 'VNM',
    'BIH': 'BIH',
    'BOL': 'BOL',
    'ESA': 'SLV', 'SLV': 'SLV',
    'PAK': 'PAK',
    'MLT': 'MLT',
    'MGL': 'MNG', 'MNG': 'MNG',
    'TUN': 'TUN',
    'MKD': 'MKD',
    'NGR': 'NGA', 'NGA': 'NGA',
    'ARM': 'ARM',
    'MDA': 'MDA',
    'MDV': 'MDV',
    'LIE': 'LIE',
    'NCA': 'NIC', 'NIC': 'NIC',
    'LIB': 'LBN', 'LBN': 'LBN',
    'UGA': 'UGA',
    'MAC': 'MAC',
    'ZIM': 'ZWE', 'ZWE': 'ZWE',
    'JAM': 'JAM',
    'CUB': 'CUB',
    'AFG': 'AFG',
    'NEP': 'NPL', 'NPL': 'NPL',
    'TAN': 'TZA', 'TZA': 'TZA',
    'AND': 'AND',
    'SWZ': 'SWZ', 'SZE': 'SWZ',
    'DJI': 'DJI',
    'GRL': 'GRL',
    'GLP': 'GLP',
    'MYA': 'MMR', 'MMR': 'MMR',
    'FIJ': 'FIJ',
    'LBY': 'LBY',
    'LCA': 'LCA',
    'UAE': 'UAE',
    'KWT': 'KWT',
    'PLE': 'PSE', 'PAL': 'PSE',
    'ANG': 'AGO',
    'SRI': 'LKA', 'LKA': 'LKA'
}

# Function to standardize country codes
def standardize_country_code(code):
    return country_mapping_cleaning.get(code, code)


def time_to_seconds(time_str):
    """
    Convert a time string in the format HH:MM:SS to seconds.
    
    Parameters:
        time_str (str): Time string in the format HH:MM:SS
    
    Returns:
        int: Time in seconds
    """
    try:
        hours, minutes, seconds = map(int, time_str.split(':'))
        return hours * 3600 + minutes * 60 + seconds
    except ValueError:
        return np.nan


def clean_and_format_name(name):
    if not isinstance(name, str):
        return name  # Return the value as is if it's not a string (e.g., NaN)
    
    # Remove country code in parentheses
    name_cleaned = re.sub(r'\s*\(\w{3}\)$', '', name)
    # Convert to title case
    name_formatted = name_cleaned.title()
    return name_formatted


def convert_seconds_to_hhmmss(seconds):
    # Ensure seconds are converted to an integer
    seconds = int(seconds)
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    seconds = seconds % 60
    return f"{hours:02}:{minutes:02}:{seconds:02}"


def get_results(df_all, time_column='tiempo_secs', genders=['M', 'F'], gender_column='sexo'):
    df_pace = df_all[(df_all[gender_column]==genders[0])]
    df_masc = df_pace[time_column].describe(percentiles=[0.01, 0.03, 0.05, 0.1, 0.2, 0.3, 0.4, 0.5, 0.75, 0.9]).drop(['std']).apply(convert_seconds_to_hhmmss)
    df_masc.iloc[0] = time_to_seconds(df_masc.iloc[0])

    df_pace = df_all[(df_all[gender_column]==genders[1])]
    df_fem = df_pace[time_column].describe(percentiles=[0.01, 0.03, 0.05, 0.1, 0.2, 0.3, 0.4, 0.5, 0.75,  0.9]).drop(['std']).apply(convert_seconds_to_hhmmss)
    df_fem.iloc[0] = time_to_seconds(df_fem.iloc[0])

    df_both = pd.concat([df_fem, df_masc], axis=1).loc[['min', '1%', '3%', '5%', '10%', '20%', '30%', '40%', '50%', '75%', '90%', 'max', 'mean', 'count']]
    df_both.columns = ['Mujeres', 'Hombres']
    
    df_both = df_both.reset_index()
    df_both.loc[df_both['index']=='min', 'index'] = 'Mínimo'
    df_both.loc[df_both['index']=='max', 'index'] = 'Máximo'
    df_both.loc[df_both['index']=='mean', 'index'] = 'Promedio'
    df_both.loc[df_both['index']=='count', 'index'] = 'Corredores'
    df_both.columns = ['Percentil', 'Mujeres', 'Hombres']

    return df_both


def get_results_distance_teams(df_all, time_column='tiempo_secs', genders=['MEN', 'WOMEN', 'MIXED'], gender_column='sexo'):
    df_pace = df_all[(df_all[gender_column]==genders[0])]
    df_masc = df_pace[time_column].describe(percentiles=[0.01, 0.02, 0.03, 0.05, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]).drop(['std']).apply(convert_seconds_to_hhmmss)
    df_masc.iloc[0] = time_to_seconds(df_masc.iloc[0])

    df_pace = df_all[(df_all[gender_column]==genders[1])]
    df_fem = df_pace[time_column].describe(percentiles=[0.01, 0.02,  0.03, 0.05, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]).drop(['std']).apply(convert_seconds_to_hhmmss)
    df_fem.iloc[0] = time_to_seconds(df_fem.iloc[0])

    df_pace = df_all[(df_all[gender_column]==genders[2])]
    df_mixed = df_pace[time_column].describe(percentiles=[0.01, 0.02,  0.03, 0.05, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]).drop(['std']).apply(convert_seconds_to_hhmmss)
    df_mixed.iloc[0] = time_to_seconds(df_mixed.iloc[0])

    df_both = pd.concat([df_fem, df_masc, df_mixed], axis=1).loc[['min', '1%', '2%', '3%', '5%', '10%', '20%', '30%', '40%', '50%', '60%', '70%', '80%', '90%', 'max', 'mean', 'count']]
    df_both.columns = ['Mujeres', 'Hombres', 'Mixto']
    
    df_both = df_both.reset_index()
    df_both.loc[df_both['index']=='min', 'index'] = 'Mínimo'
    df_both.loc[df_both['index']=='max', 'index'] = 'Máximo'
    df_both.loc[df_both['index']=='mean', 'index'] = 'Promedio'
    df_both.loc[df_both['index']=='count', 'index'] = 'Corredores'
    df_both.columns = ['Percentil', 'Mujeres', 'Hombres', 'Mixto']

    return df_both



def get_results_distance(df_all, time_column='tiempo_secs', genders=['M', 'F'], gender_column='sexo'):
    df_pace = df_all[(df_all[gender_column]==genders[0])]
    df_masc = df_pace[time_column].describe(percentiles=[0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 0.95, 0.99]).drop(['std'])

    df_pace = df_all[(df_all[gender_column]==genders[1])]
    df_fem = df_pace[time_column].describe(percentiles=[0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 0.95, 0.99]).drop(['std'])

    df_both = pd.concat([df_fem, df_masc], axis=1).loc[['min', '10%', '20%', '30%', '40%', '50%', '60%', '70%', '80%', '90%', '95%', '99%', 'max', 'mean', 'count']]
    df_both.columns = ['Mujeres', 'Hombres']
    
    df_both = df_both.reset_index()
    df_both.loc[df_both['index']=='min', 'index'] = 'Mínimo'
    df_both.loc[df_both['index']=='max', 'index'] = 'Máximo'
    df_both.loc[df_both['index']=='mean', 'index'] = 'Promedio'
    df_both.loc[df_both['index']=='count', 'index'] = 'Corredores'
    df_both.columns = ['Percentil', 'Mujeres', 'Hombres']
    df_both['Mujeres'] = df_both['Mujeres'].round(1)
    df_both['Hombres'] = df_both['Hombres'].round(1)

    return df_both


def get_results_distance_teams_12h(df_all, time_column='tiempo_secs', genders=['MEN', 'WOMEN', 'MIXED'], gender_column='sexo'):
    df_pace = df_all[(df_all[gender_column]==genders[0])]
    df_masc = df_pace[time_column].describe(percentiles=[0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 0.95, 0.99]).drop(['std'])

    df_pace = df_all[(df_all[gender_column]==genders[1])]
    df_fem = df_pace[time_column].describe(percentiles=[0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 0.95, 0.99]).drop(['std'])

    df_pace = df_all[(df_all[gender_column]==genders[2])]
    df_mixed = df_pace[time_column].describe(percentiles=[0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 0.95, 0.99]).drop(['std'])

    df_both = pd.concat([df_fem, df_masc, df_mixed], axis=1).loc[['min', '10%', '20%', '30%', '40%', '50%', '60%', '70%', '80%', '90%', '95%', '99%', 'max', 'mean', 'count']]
    df_both.columns = ['Mujeres', 'Hombres', 'Mixto']
    
    df_both = df_both.reset_index()
    df_both.loc[df_both['index']=='min', 'index'] = 'Mínimo'
    df_both.loc[df_both['index']=='max', 'index'] = 'Máximo'
    df_both.loc[df_both['index']=='mean', 'index'] = 'Promedio'
    df_both.loc[df_both['index']=='count', 'index'] = 'Corredores'
    df_both.columns = ['Percentil', 'Mujeres', 'Hombres', 'Mixto']
    df_both['Mujeres'] = df_both['Mujeres'].round(1)
    df_both['Hombres'] = df_both['Hombres'].round(1)
    df_both['Mixto'] = df_both['Mixto'].round(1)

    return df_both


def best_country(df, time_column, gender_col, country_col, countries):
    df_latam = df[(df[country_col].isin(countries))].copy()

    df_times_latam = df_latam.groupby([country_col, gender_col])[time_column].min().reset_index()

    df_times_latam['min_time'] = df_times_latam.groupby([gender_col])[time_column].transform('min')
    df_best_latam = df_times_latam[df_times_latam['min_time']==df_times_latam[time_column]]

    df_merged = pd.merge(left=df_latam,
                         right=df_best_latam[['min_time', gender_col]],
                         on=[gender_col])

    df_final_times = df_merged[df_merged['min_time']==df_merged[time_column]]
    return df_final_times

def get_top_n(df, gender_col, time_column, position):
    df['time_stamp'] = pd.to_timedelta(df[time_column])
    df['gender_position'] = df.groupby([gender_col])['time_stamp'].rank(method='min').astype('int64')
    
    return df[df['gender_position']<=position]


def get_top_n_reverse(df, gender_col, time_column, position):
    df[time_column] = df[time_column].astype(float)
    df['gender_position'] = df.groupby([gender_col])[time_column].rank(method='min', ascending=False).astype('int64')
    
    return df[df['gender_position']<=position]


def averages_country(df, time_column, country_col, apply_country_map=True, min_participants=10):

    if apply_country_map:
        df['country_name'] = df[country_col].apply(lambda x: country_map[x])
    else:
        df['country_name'] = df[country_col]

    df['country_count'] = df.groupby('country_name')['country_name'].transform('count')

    df_times = df[df['country_count']>=min_participants].groupby('country_name')[time_column].mean().apply(convert_seconds_to_hhmmss)
    df_count = df[df['country_count']>=min_participants].groupby('country_name')['country_name'].count()

    return pd.concat([df_times, df_count], axis=1).rename(columns={'country_name': 'Corredores',
                                                                   time_column: 'Tiempo'}).sort_values('Corredores', ascending=False)


def averages_country_boston(df, time_column, country_col, apply_country_map=True, min_participants=10):

    df['country_name'] = df[country_col]
    df['country_count'] = df.groupby('country_name')['country_name'].transform('count')

    df_times_M = df[(df['country_count']>=min_participants) & (df['gender']=='M')].groupby('country_name')[time_column].mean().apply(convert_seconds_to_hhmmss)
    df_times_W = df[(df['country_count']>=min_participants) & (df['gender']=='W')].groupby('country_name')[time_column].mean().apply(convert_seconds_to_hhmmss)
    df_count_M = df[(df['country_count']>=min_participants) & (df['gender']=='M')].groupby('country_name')['country_name'].count()
    df_count_W = df[(df['country_count']>=min_participants) & (df['gender']=='W')].groupby('country_name')['country_name'].count()
    df_boston_M = df[(df['country_count']>=min_participants) & (df['gender']=='M')].groupby('country_name')['boston_qualified'].sum()
    df_boston_W = df[(df['country_count']>=min_participants) & (df['gender']=='W')].groupby('country_name')['boston_qualified'].sum()

    df_res = pd.concat([df_times_W, df_count_W, df_boston_W, df_times_M, df_count_M, df_boston_M], axis=1)
    df_res.columns = ['Tiempo promedio (W)', 'Corredores (W)', 'BQs (W)', 
                     'Tiempo promedio', 'Corredores', 'BQs']
    df_res = df_res.sort_values('Corredores (W)', ascending=False)
    return df_res


def averages_country_gender(df, time_column, country_col, apply_country_map=True, min_participants=10):

    df['country_name'] = df[country_col]
    df['country_count'] = df.groupby('country_name')['country_name'].transform('count')

    df_times_M = df[(df['country_count']>=min_participants) & (df['gender']=='M')].groupby('country_name')[time_column].mean().apply(convert_seconds_to_hhmmss)
    df_times_W = df[(df['country_count']>=min_participants) & (df['gender']=='W')].groupby('country_name')[time_column].mean().apply(convert_seconds_to_hhmmss)
    df_count_M = df[(df['country_count']>=min_participants) & (df['gender']=='M')].groupby('country_name')['country_name'].count()
    df_count_W = df[(df['country_count']>=min_participants) & (df['gender']=='W')].groupby('country_name')['country_name'].count()

    df_res = pd.concat([df_times_W, df_count_W, df_times_M, df_count_M], axis=1)
    df_res.columns = ['Tiempo promedio (W)', 'Corredores (W)', 
                     'Tiempo promedio', 'Corredores']
    df_res = df_res.sort_values('Corredores (W)', ascending=False)
    return df_res

def count_per_interval(df, gender_col, time_col):

    # Step 1: Convert time from seconds to hours
    df['time_in_hours'] = df[time_col] / 3600

    # Step 2: Define bins (2:45:00 to 6:00:00, with 15-minute intervals)
    start_time = 2 + 30 / 60  # 2:45 in hours
    end_time = 6               # 6:00 in hours

    # Create bins (2:45:00 to 6:00:00, 15 minutes intervals)
    bins = [2]+[start_time + i * 0.25 for i in range(int((end_time - start_time) / 0.25) + 1)]+[9]

    # Format the labels as "HH:MM:SS - HH:MM:SS"
    bin_labels = [
        convert_seconds_to_hhmmss(bins[i]*3600) + "  -  " + convert_seconds_to_hhmmss(bins[i+1]*3600 - 1/3600) for i in range(len(bins) - 1)
    ]

    # Step 3: Bin the times and group by the bins and gender
    df['time_bin'] = pd.cut(df['time_in_hours'], bins=bins, labels=bin_labels, include_lowest=True, right=False)

    # Step 4: Count the number of people in each bin by gender
    result = df.groupby(['time_bin', gender_col]).size().unstack(fill_value=0)

    return result


def count_boston_qualifiers(df, time_col, cat_col, gender_col):

    df['boston_threshold'] = df.apply(lambda x: boston_qualifying_times[x[cat_col]][x[gender_col]], axis=1).apply(time_to_seconds)
    df['boston_qualifier'] = df[time_col]<=df['boston_threshold']
    boston_stats = df.groupby([cat_col, gender_col])[['boston_qualifier', 'average']].agg({'boston_qualifier': 'sum',
                                                                                        'average': 'count'}).reset_index()
    boston_stats = boston_stats.rename(columns={cat_col: 'Categoría', 'boston_qualifier': 'Clasificados', gender_col: 'Género', 'average': 'Corredores'})
    boston_stats = boston_stats[boston_stats['Categoría']!='\x13']
    if len(boston_stats[boston_stats['Categoría']=='JU20'])>0:
        reorder = 1
    else:
        reorder = 0
    boston_stats['Categoría'] = boston_stats.apply(lambda x: categories[x['Categoría']], axis=1)
    boston_stats=boston_stats.sort_values('Categoría')
    boston_stats= boston_stats.pivot(index='Categoría', columns='Género', values=['Clasificados', 'Corredores']).reset_index().fillna(0)
    boston_stats.columns = boston_stats.columns.swaplevel(0, 1)
    boston_stats = boston_stats.sort_index(axis=1)
    if reorder==1:
        boston_stats = pd.concat([boston_stats.iloc[[-1]], boston_stats.iloc[:-1]]).reset_index(drop=True)
    return boston_stats



def save_averages_country_file():
    # Create a Pandas Excel writer using XlsxWriter as the engine
    writer = pd.ExcelWriter("berlin_analysis/best_latam_countries.xlsx", engine='xlsxwriter')

    df_countries = averages_country(df[df['Country'].isin(latam)], time_column, country_col).sort_values('Tiempo')
    df_countries = df_countries.reset_index().rename(columns={'country_name': 'País'})

    # Write the dataframe to Excel
    df_countries.to_excel(writer, sheet_name='Sheet1', index=False)

    # Access the workbook and worksheet objects
    workbook  = writer.book
    worksheet = writer.sheets['Sheet1']

    # Define the format for headers with font size 18, bold, and borders
    header_format = workbook.add_format({
        'font_size': 18,
        'bold': True,
        'border': 1  # Adds a thin border only to the headers
    })

    # Define the format for cells with font size 18 (no borders)
    cell_format = workbook.add_format({
        'font_size': 18
    })

    # Apply the format to the header row (row 0, starting from column A)
    for col_num, value in enumerate(df_countries.columns.values):
        worksheet.write(0, col_num, value, header_format)

    # Apply the format to the entire DataFrame (without borders)
    for row_num, row_data in df_countries.iterrows():
        for col_num, cell_data in enumerate(row_data):
            worksheet.write(row_num + 1, col_num, cell_data, cell_format)

    # Set the column widths to automatically fit the content
    for col_num, col in enumerate(df_countries.columns):
        max_len = max(df_countries[col].astype(str).map(len).max(), len(col))  # Find max content width
        worksheet.set_column(col_num, col_num, max_len + 2)  # Adjust the column width

    # Save the Excel file
    writer._save()
    
    

def plot_distance(df, time_column, save_path):
    plt.rcParams.update({'font.size': 18})

    # Create the histogram
    plt.figure(figsize=(12, 5))
    plt.hist(
        df[time_column], bins=30, color='purple', alpha=0.3, edgecolor='black', label='All Runners'
    )

    plt.xlabel("Distancia", fontsize=18)   # Set font size for x-axis label
    plt.ylabel("Corredores", fontsize=18)  # Set font size for y-axis label

    plt.xticks(fontsize=16)  # Set font size for x-axis ticks
    plt.yticks(fontsize=16)  # Set font size for y-axis ticks

    #plt.legend(loc='upper center', bbox_to_anchor=(0.5, -0.25), fontsize=18, ncol=2)  # Place legend below the plot
    plt.tight_layout()  # Adjust layout to fit everything nicely
    plt.savefig(save_path)
    plt.close()
    

def top_n_save_dataframe_to_excel(df, file_path):
    """
    Saves a DataFrame to an Excel file with specific formatting:
    - First column left-aligned, others center-aligned
    - Header row bold with light gray background and wrapped text
    - Auto-fit column width applied at the end
    - Font size 18, Helvetica Neue (or similar)
    - Borders only on the outside of the table
    """
    # Export to Excel
    df.to_excel(file_path, index=False, engine="openpyxl")

    # Load workbook and select active sheet
    wb = load_workbook(file_path)
    ws = wb.active

    # Define styles
    font_style = Font(name="Helvetica Neue", size=18)  # Font for all cells
    header_font = Font(name="Helvetica Neue", size=18, bold=True)  # Bold for headers
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Light gray background
    thin = Side(style="thin")
    wrap_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # Wrap text for headers

    # Borders (only applied to the outer edges)
    border_top = Border(top=thin)
    border_bottom = Border(bottom=thin)
    border_left = Border(left=thin)
    border_right = Border(right=thin)

    # Get the dimensions of the table
    max_row = ws.max_row
    max_col = ws.max_column

    # Dictionary to track max content width per column
    column_widths = {}

    # Apply formatting
    for col_idx, col in enumerate(ws.iter_cols(), start=1):
        for row_idx, cell in enumerate(col, start=1):
            cell.font = font_style  # Apply font
            if row_idx == 1:  # Header row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = wrap_alignment  # Apply wrapped text for headers
            else:
                cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "center")

            # Apply border only to the outer edges of the table
            if row_idx == 1:
                cell.border = Border(top=thin, bottom=thin)  # Top border for the first row
            if row_idx == max_row:
                cell.border = Border(bottom=thin)  # Bottom border for the last row
            if col_idx == 1:
                cell.border = Border(left=thin)  # Left border for the first column
            if col_idx == max_col:
                cell.border = Border(right=thin)  # Right border for the last column

            if (row_idx == 1) & (col_idx == 1):
                cell.border = Border(top=thin, bottom=thin, left=thin)  # Top border for the first row
            if (row_idx == max_row) & (col_idx == 1):
                cell.border = Border(bottom=thin, left=thin)  # Bottom border for the last row
            if (row_idx == 1) & (col_idx == max_col):
                cell.border = Border(top=thin, bottom=thin, right=thin)  # Top border for the first row
            if (row_idx == max_row) & (col_idx == max_col):
                cell.border = Border(bottom=thin, right=thin)  # Bottom border for the last row
                
            # Track max content length for column width adjustment
            if cell.value:
                column_widths[col_idx] = max(column_widths.get(col_idx, 0), len(str(cell.value)))

    # Apply final column width adjustments **after** all styling
    for col_idx, width in column_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width + 5  # Adjust for better fit

    # Save the modified Excel file
    wb.save(file_path)
    
    
def distribution_save_dataframe_to_excel_centered(df, file_path):
    """
    Saves a DataFrame to an Excel file with specific formatting:
    - All columns center-aligned
    - Header row bold with light gray background
    - Auto-fit column width applied at the end
    - Font size 18, Helvetica Neue (or similar)
    - Borders only on the outside of the table
    """
    # Export to Excel
    df.to_excel(file_path, index=False, engine="openpyxl")

    # Load workbook and select active sheet
    wb = load_workbook(file_path)
    ws = wb.active

    # Define styles
    font_style = Font(name="Helvetica Neue", size=18)  # Font for all cells
    header_font = Font(name="Helvetica Neue", size=18, bold=True)  # Bold for headers
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Light gray background
    wrap_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # Centered + wrap text for headers

    # Borders (only applied to the outer edges)
    thin = Side(style="thin")
    border_top = Border(top=thin)
    border_bottom = Border(bottom=thin)
    border_left = Border(left=thin)
    border_right = Border(right=thin)

    # Get the dimensions of the table
    max_row = ws.max_row
    max_col = ws.max_column

    # Dictionary to track max content width per column
    column_widths = {}

    # Apply formatting
    for col_idx, col in enumerate(ws.iter_cols(), start=1):
        for row_idx, cell in enumerate(col, start=1):
            cell.font = font_style  # Apply font
            cell.alignment = Alignment(horizontal="center")  # Center align all cells

            if row_idx == 1:  # Header row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = wrap_alignment  # Wrap text for headers

            if row_idx == max_row-1:  # Header row
                cell.font = header_font
                cell.fill = header_fill

            # Apply border only to the outer edges of the table
            if row_idx == 1:
                cell.border = Border(top=thin, bottom=thin)  # Top border for the first row
            if row_idx == max_row:
                cell.border = Border(bottom=thin)  # Bottom border for the last row
            if row_idx == max_row-1:
                cell.border = Border(top=thin, bottom=thin)  # Bottom border for the last row
            if col_idx == 1:
                cell.border = Border(left=thin)  # Left border for the first column
            if col_idx == max_col:
                cell.border = Border(right=thin)  # Right border for the last column

            if (row_idx == max_row-1) & (col_idx == 1):
                cell.border = Border(top=thin, bottom=thin, left=thin)  # Top border for the first row
            if (row_idx == max_row-1) & (col_idx == max_col):
                cell.border = Border(top=thin, bottom=thin, right=thin)  # Top border for the first row

            if (row_idx == 1) & (col_idx == 1):
                cell.border = Border(top=thin, bottom=thin, left=thin)  # Top border for the first row
            if (row_idx == max_row) & (col_idx == 1):
                cell.border = Border(bottom=thin, left=thin)  # Bottom border for the last row
            if (row_idx == 1) & (col_idx == max_col):
                cell.border = Border(top=thin, bottom=thin, right=thin)  # Top border for the first row
            if (row_idx == max_row) & (col_idx == max_col):
                cell.border = Border(bottom=thin, right=thin)  # Bottom border for the last row

            # Track max content length for column width adjustment
            if cell.value:
                column_widths[col_idx] = max(column_widths.get(col_idx, 0), len(str(cell.value)))

    # Apply final column width adjustments **after** all styling
    for col_idx, width in column_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width + 10  # Adjust for better fit

    # Save the modified Excel file
    wb.save(file_path)
    
    

def apply_opacity(image, opacity):
    image = image.convert("RGBA")
    r, g, b, a = image.split()
    # Scale alpha channel by the desired opacity
    a = a.point(lambda p: int(p * (opacity / 255)))
    return Image.merge("RGBA", (r, g, b, a))


def render_table_to_image(df, font_size=18, row_height=0.12, table_opacity=255, table_scale=0.5):
    fig, ax = plt.subplots()
    ax.axis('off')

    table = ax.table(cellText=df.values,
                     colLabels=df.columns,
                     cellLoc='center',
                     loc='center')

    table.auto_set_font_size(False)
    table.set_fontsize(font_size)

    for (row, col), cell in table.get_celld().items():
        cell.set_height(row_height)
        if (row == 0) | (row==13):
            cell.set_fontsize(font_size)
            cell.set_text_props(weight='bold')
            cell.set_facecolor('#D9D9D9')
        cell.set_edgecolor('black')
        cell.set_linewidth(1)

    buf = io.BytesIO()
    fig.savefig(buf, format='png', bbox_inches='tight', dpi=300, transparent=True)
    plt.close(fig)
    buf.seek(0)
    
    table_img = Image.open(buf)
    table_img = resize_image(table_img, scale=table_scale)

    return apply_opacity(table_img, table_opacity)



from PIL import Image, ImageOps


def plot_histogram_with_save(df, time_column, save_path, start_time=2*3600, end_time=6*3600, bin_width=15*60, y_title='Número de corredores'):
    """
    Creates and saves a histogram of finish times.
    
    Parameters:
        df (pd.DataFrame): The DataFrame containing the finish times.
        time_column (str): The column name containing times in seconds.
        save_path (str): The path to save the plot (including filename and extension).
    """
    # Define the range and new bin width
    new_bin_width = bin_width // 3  # Divide bin width by 3
    bins = np.arange(start_time, end_time + new_bin_width, new_bin_width)

    # Combine all finish times
    all_times = df[time_column]

    # Set global font size
    plt.rcParams.update({'font.size': 18})

    # Create the histogram
    plt.figure(figsize=(8, 6.8), facecolor='white')
    plt.hist(
        all_times, bins=bins, color='purple', alpha=0.3, edgecolor='black', label='All Runners'
    )

    # Format the x-axis ticks (same number as original)
    tick_positions = np.arange(start_time, end_time + bin_width, bin_width)
    tick_labels = [convert_seconds_to_hhmmss(pos) for pos in tick_positions]
    plt.xticks(tick_positions, tick_labels, rotation=45, ha='right')

    # Set limits for x-axis
    plt.xlim(start_time, end_time)

    # Add labels and title
    plt.xlabel('Tiempo total (HH:MM:SS)', fontsize=18)
    plt.ylabel(y_title, fontsize=18)

    # Add black border around the entire plot
    ax = plt.gca()
    for spine in ax.spines.values():
        spine.set_edgecolor('black')
        spine.set_linewidth(2)  # Grosor del borde
    
    # Save plot
    plt.tight_layout()
    plt.savefig(save_path, facecolor='white')
    plt.close()  # Close the plot to free memory
    
    # Add black border to the saved image
    border_size = 2  # pixels
    with Image.open(save_path) as im:
        bordered = ImageOps.expand(im, border=border_size, fill='black')
        bordered.save(save_path)



def add_texts(text, x_vert, y_vert, draw, background, fontsize=40, angle=0, centered=False, tag=False, tag_color="white", padding=10, radius=10, margin=10, font_type="HelveticaNeueMedium.otf"):
    
    font = ImageFont.truetype(font_type, fontsize)
    text_width, text_height = draw.textsize(text, font=font)
    if tag:
        # Total box size
        box_width = text_width + 2 * padding
        box_height = text_height + 2 * padding
        
        x_vert=background.width - box_width - margin
        y_vert=background.height - box_height - margin
            
        # Draw white rounded rectangle
        draw.rounded_rectangle(
            [(x_vert, y_vert), (x_vert + box_width, y_vert + box_height)],
            radius=radius,
            fill=tag_color
        )

        # Draw black text centered inside box
        x_text = x_vert + padding
        y_text = y_vert + padding
        draw.text((x_text, y_text), text, font=font, fill="black")

    else:

        if centered:
            x_vert = (background.width - text_width) // 2
        # Create a new transparent image for the text
        text_size = draw.textsize(text, font=font)
        temp_img = Image.new("RGBA", (text_size[0], text_size[1]), (255, 255, 255, 0))
        temp_draw = ImageDraw.Draw(temp_img)
        temp_draw.text((0, 0), text, font=font, fill="black")

        # Rotate 90° for vertical text (top-to-bottom)
        rotated_text_img = temp_img.rotate(angle, expand=1)

        # Paste rotated text onto the background
        background.paste(rotated_text_img, (x_vert, y_vert), rotated_text_img)


def smart_wrap(text, max_len=100):
    return text[:(max_len-15)]


def render_top_n_table_to_image(df, top_n=10, font_size=18,
                                row_height=0.12, col_widths=None,
                                table_opacity=255, name_col_width=1.2,
                                name_col_max_length=50, table_scale=0.5,
                                indicator_column=None,
                                country_col=None):

    fig, ax = plt.subplots()
    ax.axis('off')

    col_align = ['left'] + ['center'] * (len(df.columns) - 1)

    df_wrapped = df.copy().astype(str).applymap(lambda x: smart_wrap(x, max_len=name_col_max_length))

    if indicator_column:
        df['previous'] = (
            df['previous']
            .replace('', 0)       # reemplaza cadenas vacías por NA
            .astype('Int64')          # convierte a entero con soporte para nulos
        )
        
        highlight_mask = df[indicator_column].astype(int) == 0
        df_wrapped = df_wrapped.drop(columns=[indicator_column])
        
        col_align = ['left'] + ['center'] * (len(df_wrapped.columns) - 1)
        col_labels = [smart_wrap(col, max_len=15) for col in df_wrapped.columns]
        col_labels = ['\n'.join(col.split()) for col in df_wrapped.columns]
    else:
        highlight_mask = pd.Series([False] * len(df_wrapped))
        col_labels = [smart_wrap(col, max_len=15) for col in df.columns]
        col_labels = ['\n'.join(col.split()) for col in df.columns]

    table = ax.table(
        cellText=df_wrapped.values,
        colLabels=col_labels,
        cellLoc='center',
        loc='center',
        colColours=['#D9D9D9'] * len(df.columns)
    )

    table.auto_set_font_size(False)
    table.set_fontsize(font_size)

    for (row, col), cell in table.get_celld().items():
        # Set height — double it for header row
        if row == 0:
            cell.set_height(row_height * 2)
            cell.set_linewidth(1)
            cell.set_text_props(weight='bold', ha=col_align[col], va='center')
            cell.set_facecolor('#D9D9D9')
        else:
            cell.set_height(row_height)
            cell.set_text_props(ha=col_align[col], va='center')
            
            data_row_idx = row - 1
            if row > 0:  # only apply to data rows
                if highlight_mask.iloc[data_row_idx]:
                    cell.set_facecolor('#FFF59D')  # yellow
                else:
                    cell.set_facecolor('white')

        if (row == (top_n+2)) | (row == 1):
            cell.visible_edges = 'B'
            cell.set_facecolor('#D9D9D9')
            cell.set_height(0.05)
            
        cell.set_edgecolor('black')

        if col == 0:
            cell.set_width(name_col_width)
        else:
            cell.set_width(0.3)
            
        if country_col != None:
            if col == country_col[0]:
                cell.set_width(country_col[1])


    # Save to buffer
    buf = io.BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight', dpi=300, transparent=True)
    plt.close(fig)
    buf.seek(0)

    # Resize to 50%
    table_img = Image.open(buf)
    table_img = resize_image(table_img, scale=table_scale)

    return apply_opacity(table_img, table_opacity)


def resize_image(img: Image.Image, scale: float) -> Image.Image:
    """
    Resize an image by a percentage scale.

    Parameters:
    - img (PIL.Image): The image to resize.
    - scale (float): Percentage to scale (e.g., 0.5 for 50%).

    Returns:
    - Resized PIL.Image.
    """
    new_width = max(1, int(img.width * scale))
    new_height = max(1, int(img.height * scale))
    return img.resize((new_width, new_height), resample=Image.LANCZOS)


def image_resize(table_img, x_table, y_table, downscaling, background, centered=False):

    table_img = table_img.resize(
        (table_img.width // int((1/downscaling)), table_img.height // int(1/downscaling)),
        resample=Image.LANCZOS  # high-quality downscaling
    )

    if centered:
        x_table = (background.width - table_img.width) // 2

        background.paste(
            table_img,
            (x_table, y_table),
            mask=table_img if table_img.mode == "RGBA" else None
        )
        
    else:
        background.paste(table_img, (x_table, y_table), mask=table_img if table_img.mode == "RGBA" else None)
    

def insert_histogram(hist_img, x_hist, y_hist, background, target_width = 900):
    # Optional: Resize histogram to fit your layout width
    
    if hist_img.width > target_width:
        ratio = target_width / hist_img.width
        hist_img = hist_img.resize((target_width, int(hist_img.height * ratio)), Image.LANCZOS)

    # Paste with alpha mask if needed
    background.paste(hist_img, (x_hist, y_hist), mask=hist_img)

    
def load_flag(country, background, x, y):
    fixed_width = 80  # Set the desired fixed width

    # Load and resize flag
    flag = Image.open(f"flags/{country.lower()}_flag.png").convert("RGBA")
    aspect_ratio = flag.height / flag.width
    flag = flag.resize((fixed_width, int(fixed_width * aspect_ratio)), Image.LANCZOS)

    # Paste left flag
    background.paste(flag, (x, y), mask=flag)

    # Paste right flag (mirrored horizontally)
    background.paste(flag, (1000 - x, y), mask=flag)
    
    
def get_data_results(df_country, folder, gender_col, time_column, histogram_thresholds,
                     name_col, place_col, cols_top, cols_top_names, top_n=10):
    df_country = df_country[df_country[gender_col].isin(['M', 'W'])]
    df_dist = get_results(df_country,
                          time_column=time_column,
                          genders=['M', 'W'],
                          gender_column=gender_col)

    plot_histogram_with_save(df_country, time_column,
                            f"{folder}/distribution.png",
                             start_time=histogram_thresholds[0]*3600, 
                             end_time=histogram_thresholds[1]*3600)

    df_country[name_col] = df_country[name_col].apply(clean_and_format_name)

    df_top = get_top_n(df_country, gender_col, time_column, position=top_n)
    
    if place_col in cols_top:
        df_top = df_top[cols_top + [gender_col]].sort_values([gender_col, place_col], ascending=[False, True])
    else:
        df_top = df_top[cols_top + [gender_col, place_col]].sort_values([gender_col, place_col], ascending=[False, True])
        
    df_top = df_top[cols_top]
    df_top.columns = cols_top_names

    # Create an empty row with the same columns
    empty_row = pd.DataFrame([[""] * len(df_top.columns)], columns=df_top.columns)

    # Insert it after the first row (index 0)
    df_top = pd.concat([df_top.iloc[:0], empty_row, df_top.iloc[0:]], ignore_index=True)

    num_women = min(df_country.value_counts(gender_col)['W'], top_n)

    empty_rows = top_n - num_women
    df_output = df_top.iloc[0:(num_women+1)]

    for i in range(empty_rows):
        df_output = pd.concat([df_output, empty_row], ignore_index=True)

    df_output = pd.concat([df_output, empty_row, df_top.iloc[(num_women+1):]], ignore_index=True)
    
    return df_dist, df_output


def create_intagram_post(country, df_dist, df_top, folder,
                         image, map_country=True, top_n=10,
                         load_flags=True, transparency=100,
                         name_col_width=1.2, name_col_max_length=50,
                         table_scale=0.5, indicator_column=None,
                         y_vert_hombres=620, y_vert_mujeres=280,
                         country_col=None):
    
    # Step 1: Load and prep background image
    background = Image.open(f"{folder}/{image}").resize((1080, 1350)).convert("RGBA")
    
    if transparency < 100:
        white_bg = Image.new("RGBA", background.size, (255, 255, 255, 255))
        background = Image.blend(white_bg, background, alpha=transparency/100)
    
    draw = ImageDraw.Draw(background)

    if map_country:
        country_mapped = country_map[country]
    else:
        country_mapped = country

    if load_flags:
        load_flag(remove_accents(country_mapped), background, 80, 50)

    # Step 3: Render table image
    table_img = render_table_to_image(df_dist, table_scale=table_scale)

    # Step 4: Paste table image onto background
    image_resize(table_img, 120, 920, 0.4, background)

    dist_img = Image.open(f"{folder}/distribution.png").convert("RGBA")
    dist_img = resize_image(dist_img, scale=table_scale*1.7)
    
    hist_img = apply_opacity(dist_img, TABLE_OPACITY)

    insert_histogram(hist_img, 510, 920, background, target_width = 500)

    # Generate the top_n table image
    top_n_img = render_top_n_table_to_image(df_top, top_n, name_col_width=name_col_width,
                                           name_col_max_length=name_col_max_length,
                                           table_scale=table_scale,
                                           indicator_column=indicator_column,
                                           country_col=country_col)


    image_resize(top_n_img, 70, 130, 0.4, background, centered=True)


    add_texts(text=country_mapped,
              x_vert=50,
              y_vert=50, 
              draw=draw,
              background=background,
              fontsize=80, angle=0, centered=True)

    # ---- Tag settings ----
    add_texts(text="@datarunnerco",
              x_vert=5,
              y_vert=5,
              draw=draw,
              background=background,
              fontsize=20, angle=0,
              tag=True,
              margin=5,
              tag_color="white")

    add_texts(text="Mujeres", x_vert=16, y_vert=y_vert_mujeres, draw=draw,
              background=background, fontsize=40, angle=90)

    add_texts(text="Hombres", x_vert=16, y_vert=y_vert_hombres, draw=draw,
              background=background, fontsize=40, angle=90)

    # Step 5: Show final image
    os.makedirs(f"{folder}/instagram", exist_ok=True)
    background.save(f"{folder}/instagram/{country}_page.png")
#    background.show()


def make_portada(df_dist, folder, image, title, include_title=True, include_stats=True, transparency=100):

    background = Image.open(f"{folder}/{image}").resize((1080, 1350)).convert("RGBA")
    if transparency < 100:
        white_bg = Image.new("RGBA", background.size, (255, 255, 255, 255))
        background = Image.blend(white_bg, background, alpha=transparency/100)
    
    draw = ImageDraw.Draw(background)
    
    
    if include_stats:
        # Step 3: Render table image
        table_img = render_table_to_image(df_dist)

        # Step 4: Paste table image onto background
        image_resize(table_img, 80, 750, 0.22, background)

        hist_img = apply_opacity(Image.open(f"{folder}/distribution.png").convert("RGBA"), TABLE_OPACITY)

        insert_histogram(hist_img, 510, 750, background, target_width = 500)

    if include_title:
        add_texts(text="Resultados",
                  x_vert=590,
                  y_vert=50, 
                  draw=draw,
                  background=background,
                  fontsize=80, angle=0, centered=True)

        add_texts(text=title,
                  x_vert=590,
                  y_vert=140, 
                  draw=draw,
                  background=background,
                  fontsize=80, angle=0, centered=True)


    # ---- Tag settings ----
    add_texts(text="@datarunnerco",
              x_vert=5,
              y_vert=5,
              draw=draw,
              background=background,
              fontsize=20, angle=0,
              tag=True,
              margin=5,
              tag_color="white")

    # Step 5: Show final image
    os.makedirs(f"{folder}/instagram", exist_ok=True)
    background.save(f"{folder}/instagram/portada.png")
    background.show()
    
    
    
def clean_abbott(df_all):
    df_all.loc[df_all['gender']=='F', 'gender'] = 'W'

    df_all['finish_secs'] = df_all['finish_time'].apply(time_to_seconds)

    df_average_times = df_all.groupby(['athlete_id'])[['finish_secs']].mean().reset_index()

    df_names = df_all[['athlete_id', 'firstname', 'lastname', 'fullname', 'nationality', 'gender']].drop_duplicates()

    df_agg = pd.merge(left=df_names,
                     right=df_average_times,
                     on='athlete_id')

    df_agg['times'] = df_agg['finish_secs'].apply(convert_seconds_to_hhmmss)
    df_agg = df_agg[df_agg['finish_secs']>7200]

    df_all.loc[df_all['event_title'].isin(['BMW Berlin Marathon', 'BMW BERLIN-MARATHON']), 'event_title'] = 'Berlin Marathon'
    df_all.loc[df_all['event_title'].isin(['Boston Marathon presented by Bank of America', 'Boston Marathon']), 'event_title'] = 'Boston Marathon'
    df_all.loc[df_all['event_title'].isin(['TCS New York City Marathon', 'TCS New York City Marathon (NYRR)']), 'event_title'] = 'New York City Marathon'
    df_all.loc[df_all['event_title'].isin(['TCS London Marathon', 'Virgin Money London Marathon']), 'event_title'] = 'London Marathon'
    df_all.loc[df_all['event_title'].isin(['Bank of America Chicago Marathon']), 'event_title'] = 'Chicago Marathon'

    # Reshaping to wide format
    wide_df = df_all.pivot_table(
        index=['athlete_id'], 
        columns=['event_title'], 
        values=['finish_time'], 
        aggfunc='first'
    )

    # Resetting index
    wide_df.reset_index(inplace=True)

    # Display the reshaped DataFrame
    wide_df.columns = [f'{col[1]}_{col[0]}' if col[1] != '' else col[0] for col in wide_df.columns]

    df_agg = pd.merge(left=df_agg, right=wide_df, on='athlete_id')
    
    df_agg['finish_secs'] = df_agg['finish_secs'].astype(float)
    df_agg['gender_position'] = df_agg.groupby(['gender'])['finish_secs'].rank(method='min', ascending=False).astype('int64')
    
    return df_agg